import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys

def total_hours(worker_id):
	worker_df = workers[worker_id]
	total_hours = 0
	for index, row in worker_df.iterrows():
		total_hours += row['Extra Hours']

	final = pd.DataFrame({
		'Worker ID': [worker_id],
		'Date': "",
		'Entry': "",
		'Lunch Start': "",
		'Lunch End': "",
		'Exit': "",
		'Extra Hours': total_hours,
		})

	workers[worker_id] = pd.concat([workers[worker_id], final], ignore_index=True)

def is_lunch_time(entry_time, lunch_start, lunch_end, exit_time):

	if lunch_start != 'Nan' and lunch_end != 'Nan':
		# Calculate difference between lunch start and end
		lunch_start_time = datetime.strptime(lunch_start, '%H:%M:%S')
		lunch_end_time = datetime.strptime(lunch_end, '%H:%M:%S')
		lunch_difference = (lunch_end_time - lunch_start_time).total_seconds()

		# See if lunch time is less than 40 minutes
		if lunch_difference / 60 < 40:
			return 0.5
		
	if entry_time.hour <= 13 and exit_time.hour > 14:
		return 1
	
	return 0

def calculate_extra_hours(entry, lunch_start, lunch_end, exit):
	extra_hours = 0
	# Convert the entry and exit times to datetime objects
	entry_time = datetime.strptime(entry, '%H:%M:%S')
	exit_time = datetime.strptime(exit, '%H:%M:%S')

	if entry_time.minute > 40:
		entry_time = entry_time.replace(hour=entry_time.hour + 1, minute=0, second=0)
	elif entry_time.minute < 20:
		entry_time = entry_time.replace(minute=0, second=0)
	else:
		entry_time = entry_time.replace(minute=30, second=0)

	if exit_time.minute > 40:
		exit_time = exit_time.replace(hour=exit_time.hour + 1, minute=0, second=0)
	elif exit_time.minute < 20:
		exit_time = exit_time.replace(minute=0, second=0)
	else:
		exit_time = exit_time.replace(minute=30, second=0)


	# Calculate the difference between the entry and exit times
	difference = (exit_time - entry_time).total_seconds() / 3600

	# Convert the difference to hours
	extra_hours = round(difference, 2) - 8 - is_lunch_time(entry_time, lunch_start, lunch_end, exit_time)

	return extra_hours

# Detects the extra hours besides the entry, lunch start, lunch end and exit times and removes them from the date and stores them in a list
def detect_incorrect_hours(date):
	incorrect_hours = []
	for i in range(1, len(date) - 2):
		time = datetime.strptime(date['tempo'][i], '%H:%M:%S')
		if time.hour < 12 or time.hour > 14:
			incorrect_hours.append(date['tempo'][i])
		date = date.drop(i)
		date = date.reset_index(drop=True)

	return date, incorrect_hours

def calculate_neg_hours(exit):
	neg_hours = 0

	exit_time = datetime.strptime(exit, '%H:%M:%S')
	if exit_time.minute > 40:
		exit_time = exit_time.replace(hour=exit_time.hour + 1, minute=0, second=0)
	elif exit_time.minute < 20:
		exit_time = exit_time.replace(minute=0, second=0)
	else:
		exit_time = exit_time.replace(minute=30, second=0)
	
	if exit_time.hour < 18:
		neg_hours = 18-exit_time.hour
		if exit_time.minute == 30:
			neg_hours -= 0.5
	return -neg_hours



def store_date(worker_id, date):
	# Convert the input date_group to a Polars DataFrame
	incorrect_hours = []
	negative_hours = 0

	if len(date) > 4:
		date, incorrect_hours = detect_incorrect_hours(date)
		entry_date = date['tempo'][0]
		lunch_start = date['tempo'][1]
		lunch_end = date['tempo'][2]
		exit_time = date['tempo'][len(date) - 1]
		extra_hours = calculate_extra_hours(entry_date, lunch_start, lunch_end, exit_time)


	elif len(date) == 1:
		entry_date = date['tempo'][0]
		lunch_start = 'Nan'
		lunch_end = 'Nan'
		exit_time = 'Nan'
		extra_hours = 0
	
	elif len(date) == 2:
		entry_date = date['tempo'][0]
		lunch_start = 'Nan'
		lunch_end = 'Nan'
		exit_time = date['tempo'][1]
		negative_hours = calculate_neg_hours(exit_time)
		extra_hours = calculate_extra_hours(entry_date, lunch_start, lunch_end, exit_time)

	
	elif len(date) == 3:

		entry_date = date['tempo'][0]
		lunch_start = date['tempo'][1]
		lunch_end = 'Nan'
		exit_time = date['tempo'][2]
		extra_hours = calculate_extra_hours(entry_date, lunch_start, lunch_end, exit_time)

	elif len(date) == 4:
		entry_date = date['tempo'][0]
		lunch_start = date['tempo'][1]
		lunch_end = date['tempo'][2]
		exit_time = date['tempo'][3]
		extra_hours = calculate_extra_hours(entry_date, lunch_start, lunch_end, exit_time)
	
	if extra_hours < 0:
		negative_hours = 0

	if incorrect_hours != []:
		incorrect_hours = ' '.join(incorrect_hours)
	else:
		incorrect_hours = ''

	new_data = pd.DataFrame({
		'Worker ID': [worker_id],
		'Date': [date['Data'][0]],
		'Entry': [entry_date],
		'Lunch Start': [lunch_start],
		'Lunch End': [lunch_end],
		'Exit': [exit_time],
		'Extra Hours': [extra_hours],
		'Negative Hours': negative_hours,
		'Incorrect Hours': [incorrect_hours]

	})

	# Check if the worker already has a DataFrame in the workers dictionary
	if worker_id in workers:
		# If the worker exists, concatenate the new data with the existing DataFrame
		workers[worker_id] = pd.concat([workers[worker_id], new_data], ignore_index=True)
	else:
		# If the worker does not exist, create a new DataFrame for the worker
		workers[worker_id] = new_data


def run(file_path):

	data = pd.read_excel(file_path)

	worker_data = data.groupby('Nome')

	global workers 
	workers = {}
	# Get unique workers
	for worker_id, worker_group in worker_data:
		date = worker_group.groupby('Data')
		for date_id, date_group in date:
			store_date(worker_id, date_group.reset_index())
		total_hours(worker_id)

	file_name = file_path.split(".")[0]
	output_path = file_name + "_result.xlsx"

	with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
		for worker_name, worker_df in workers.items():
			worker_df.to_excel(writer, sheet_name=worker_name, index=False)


			# Load the workbook
			workbook = writer.book
			worksheet = workbook[worker_name]

			# Iterate through each row and highlight rows with the specific value
			for row in range(2, len(worker_df) + 2): 
				for col in range(1, worker_df.shape[1] + 1):
					if worksheet.cell(row=row, column=col).value == 'Nan':
						worksheet[row][col - 1].fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
					if col == 7:
						if float(worksheet.cell(row=row, column=col).value) < 0:
							worksheet[row][col - 1].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
						if float(worksheet.cell(row=row, column=col).value) > 0:
							worksheet[row][col - 1].fill = PatternFill(start_color="1eff00", end_color="1eff00", fill_type="solid")
					if col == 8 and worksheet.cell(row=row, column=col).value != '':
						if float(worksheet.cell(row=row, column=col).value) < 0:
							worksheet[row][col - 1].fill = PatternFill(start_color="4287f5", end_color="ff0000", fill_type="solid")


			# Correct the column width
			for col in worksheet.columns:
				max_length = 0
				column = col[0].column_letter # Get the column name
				for cell in col:
					try: # Necessary to avoid error on empty cells
						if len(str(cell.value)) > max_length:
							max_length = len(cell.value)
					except:
						pass
				adjusted_width = (max_length)
				worksheet.column_dimensions[column].width = adjusted_width

			
			# Save the workbook
			workbook.save(output_path)
